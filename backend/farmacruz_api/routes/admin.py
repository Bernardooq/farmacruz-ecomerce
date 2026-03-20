"""
Routes de Administracion de Usuarios Internos

Endpoints CRUD para gestion de usuarios del sistema:
- GET /users - Lista de usuarios con filtros
- POST /users - Crear usuario
- GET /users/{id} - Detalle de usuario
- PUT /users/{id} - Actualizar usuario
- DELETE /users/{id} - Eliminar usuario

Permisos: Solo administradores

Nota: Este modulo gestiona usuarios INTERNOS (admin, marketing, seller).
Para clientes, ver routes/customers.py
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from typing import List, Optional
from sqlalchemy.orm import Session

from dependencies import get_db, get_current_admin_user
from db.base import User, UserRole
from crud.crud_customer import get_customer_by_email, get_customer_by_username
from schemas.user import User as UserSchema, UserUpdate, UserCreate
from crud.crud_user import (
    get_users, get_user, update_user, delete_user,
    create_user, get_user_by_username, get_user_by_email
)
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse
from datetime import datetime
from db.base import (
    Customer, CustomerInfo, SalesGroup, GroupMarketingManager, GroupSeller,
    Category, Product, PriceList, PriceListItem
)

router = APIRouter()

""" GET /users - Lista de usuarios con filtros """
@router.get("/users", response_model=List[UserSchema])
def read_all_users(
    skip: int = Query(0, ge=0, description="Registros a saltar"),
    limit: int = Query(100, ge=1, le=200, description="Maximo de registros"),
    role: Optional[str] = Query(None, description="Filtrar por rol: admin, marketing, seller"),
    search: Optional[str] = Query(None, description="Buscar por nombre o username"),
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    # Lista de usuarios internos con filtros
    role_filter = None
    if role:
        try:
            role_filter = UserRole(role)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Rol invalido: '{role}'. Roles validos: admin, marketing, seller"
            )
    
    users = get_users(db, skip=skip, limit=limit, role=role_filter, search=search)
    return users

""" POST /users - Crear usuario """
@router.post("/users", response_model=UserSchema, status_code=status.HTTP_201_CREATED)
def create_new_user(user: UserCreate, current_user: User = Depends(get_current_admin_user), db: Session = Depends(get_db)):
    db_user = get_user_by_username(db, username=user.username)
    db_customer = get_customer_by_username(db, username=user.username)
    if db_user or db_customer:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El nombre de usuario ya esta registrado"
        )
    
    # Validar email unico
    """if user.email:
        db_customer = get_customer_by_email(db, email=user.email)
        db_user = get_user_by_email(db, email=user.email)
        if db_user or db_customer:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El email ya esta registrado"
            )"""
    
    return create_user(db=db, user=user)

""" GET /users/{user_id} - Detalle de usuario """
@router.get("/users/{user_id}", response_model=UserSchema)
def read_user_by_id(user_id: int, current_user: User = Depends(get_current_admin_user), db: Session = Depends(get_db)):
    user = get_user(db, user_id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario no encontrado"
        )
    return user

""" PUT /users/{user_id} - Actualizar usuario """
@router.put("/users/{user_id}", response_model=UserSchema)
def update_user_info(user_id: int, user_update: UserUpdate, current_user: User = Depends(get_current_admin_user), db: Session = Depends(get_db)):
    
    db_customer = get_customer_by_username(db, username=user_update.username) if user_update.username else None
    # db_customer_email = get_customer_by_email(db, email=user_update.email) if user_update.email else None
    if  db_customer:
        raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="El nombre de usuario ya esta registrado"
        )
    """
    if  db_customer or db_customer_email:
        raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="El nombre de usuario o email ya esta registrado"
        )
    """
    user = update_user(db, user_id=user_id, user=user_update)
    if not user:    
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario no encontrado"
        )
    return user
    
""" DELETE /users/{user_id} - Eliminar usuario """
@router.delete("/users/{user_id}")
def delete_user_account(
    user_id: int,
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    user = delete_user(db, user_id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario no encontrado"
        )
    
    return {"message": "Usuario eliminado exitosamente"}


""" PUT /users/{user_id}/promote - Cambiar rol seller<->marketing """
@router.put("/users/{user_id}/promote")
def promote_user(
    user_id: int,
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """
    Promueve un usuario entre seller <-> marketing.
    - Si es seller -> se convierte en marketing
    - Si es marketing -> se convierte en seller
    - NUNCA se puede promover a/desde admin
    Todas las membresías de grupo se migran automáticamente.
    Solo accesible por admin.
    """
    user = db.query(User).filter(User.user_id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario no encontrado"
        )

    # Bloquear cualquier operación con admin
    if user.role == UserRole.admin:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se puede cambiar el rol de un administrador"
        )

    if user.role == UserRole.seller:
        # Seller → Marketing: mover GroupSeller → GroupMarketingManager
        seller_memberships = db.query(GroupSeller).filter(GroupSeller.seller_id == user_id).all()

        for membership in seller_memberships:
            # Verificar que no exista ya como marketing en ese grupo
            existing = db.query(GroupMarketingManager).filter(
                GroupMarketingManager.sales_group_id == membership.sales_group_id,
                GroupMarketingManager.marketing_id == user_id
            ).first()
            if not existing:
                new_membership = GroupMarketingManager(
                    sales_group_id=membership.sales_group_id,
                    marketing_id=user_id
                )
                db.add(new_membership)
            db.delete(membership)

        user.role = UserRole.marketing
        new_role = "marketing"

    elif user.role == UserRole.marketing:
        # Marketing → Seller: mover GroupMarketingManager → GroupSeller
        marketing_memberships = db.query(GroupMarketingManager).filter(
            GroupMarketingManager.marketing_id == user_id
        ).all()

        for membership in marketing_memberships:
            # Verificar que no exista ya como seller en ese grupo
            existing = db.query(GroupSeller).filter(
                GroupSeller.sales_group_id == membership.sales_group_id,
                GroupSeller.seller_id == user_id
            ).first()
            if not existing:
                new_membership = GroupSeller(
                    sales_group_id=membership.sales_group_id,
                    seller_id=user_id
                )
                db.add(new_membership)
            db.delete(membership)

        user.role = UserRole.seller
        new_role = "seller"

    db.commit()
    db.refresh(user)

    return {
        "message": f"Usuario '{user.full_name}' promovido a {new_role} exitosamente",
        "user_id": user.user_id,
        "new_role": new_role,
        "full_name": user.full_name
    }

""" GET /export-xlsx - Exportar data del sistema a Excel por tipo """
@router.get("/export-xlsx")
def export_data_xlsx(
    type: str = Query(..., description="Tipo de datos: clientes, vendedores, marketing, grupos, productos, precios"),
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """
    Exporta datos a XLSX según el tipo solicitado.
    Tipos válidos: clientes, vendedores, marketing, grupos, productos, precios
    Solo accesible por admin.
    """

    valid_types = ["clientes", "vendedores", "marketing", "grupos", "productos"]
    if type not in valid_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tipo inválido: '{type}'. Tipos válidos: {', '.join(valid_types)}"
        )

    wb = Workbook()
    ws = wb.active

    # Estilos para headers
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    def style_headers(ws, headers):
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    if type == "clientes":
        ws.title = "Clientes"
        style_headers(ws, [
            "ID", "Username", "Email", "Nombre Completo", "Activo", "Agente ID",
            "Razón Social", "RFC", "Grupo Ventas ID", "Lista Precios ID",
            "Dirección 1", "Dirección 2", "Dirección 3", "Teléfono 1", "Teléfono 2", "Creado"
        ])
        for c in db.query(Customer).outerjoin(CustomerInfo).order_by(Customer.customer_id).all():
            ci = c.customer_info
            ws.append([
                c.customer_id, c.username, c.email, c.full_name, c.is_active, c.agent_id,
                ci.business_name if ci else "", ci.rfc if ci else "",
                ci.sales_group_id if ci else "", ci.price_list_id if ci else "",
                ci.address_1 if ci else "", ci.address_2 if ci else "", ci.address_3 if ci else "",
                ci.telefono_1 if ci else "", ci.telefono_2 if ci else "",
                c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""
            ])
        filename_label = "clientes"

    elif type == "vendedores":
        ws.title = "Vendedores"
        style_headers(ws, ["ID", "Username", "Email", "Nombre Completo", "Activo", "Creado"])
        for u in db.query(User).filter(User.role == UserRole.seller).order_by(User.user_id).all():
            ws.append([
                u.user_id, u.username, u.email, u.full_name, u.is_active,
                u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else ""
            ])
        filename_label = "vendedores"

    elif type == "marketing":
        ws.title = "Marketing"
        style_headers(ws, ["ID", "Username", "Email", "Nombre Completo", "Activo", "Creado"])
        for u in db.query(User).filter(User.role == UserRole.marketing).order_by(User.user_id).all():
            ws.append([
                u.user_id, u.username, u.email, u.full_name, u.is_active,
                u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else ""
            ])
        filename_label = "marketing"

    elif type == "grupos":
        ws.title = "Grupos de Ventas"
        style_headers(ws, ["ID", "Nombre", "Descripción", "Activo"])
        for g in db.query(SalesGroup).order_by(SalesGroup.sales_group_id).all():
            ws.append([g.sales_group_id, g.group_name, g.description, g.is_active])

        # Segunda hoja: Miembros
        ws2 = wb.create_sheet("Miembros")
        style_headers(ws2, ["Grupo ID", "Grupo Nombre", "Tipo", "User ID", "Username", "Nombre Completo"])
        for gm in db.query(GroupMarketingManager).all():
            group = gm.sales_group
            user = gm.marketing_user
            ws2.append([
                group.sales_group_id if group else "", group.group_name if group else "",
                "Marketing", user.user_id if user else "", user.username if user else "",
                user.full_name if user else ""
            ])
        for gs in db.query(GroupSeller).all():
            group = gs.sales_group
            user = gs.seller_user
            ws2.append([
                group.sales_group_id if group else "", group.group_name if group else "",
                "Seller", user.user_id if user else "", user.username if user else "",
                user.full_name if user else ""
            ])
        auto_width(ws2)
        filename_label = "grupos_y_miembros"

    elif type == "productos":
        ws.title = "Productos"
        style_headers(ws, [
            "ID", "Código Barras", "Nombre", "Descripción", "Desc. 2", "Unidad",
            "Precio Base", "IVA %", "Stock", "Activo", "Categoría ID"
        ])
        for p in db.query(Product).order_by(Product.product_id).all():
            ws.append([
                p.product_id, p.codebar, p.name, p.description, p.descripcion_2,
                p.unidad_medida, float(p.base_price) if p.base_price else 0,
                float(p.iva_percentage) if p.iva_percentage else 0,
                p.stock_count, p.is_active, p.category_id
            ])
        filename_label = "inventario"

    auto_width(ws)

    # Generar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"farmacruz_{filename_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
